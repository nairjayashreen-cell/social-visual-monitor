from fastapi import FastAPI, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse
from datetime import datetime
from zoneinfo import ZoneInfo
import requests, os, json, sqlite3, uuid
import pandas as pd
import cv2
import numpy as np
from skimage.metrics import structural_similarity as ssim
from PIL import Image
import imagehash
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

# ─────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────
LOGO_STATE_FILE = "/tmp/logo_state.json"
UPLOAD_DIR      = "/tmp/uploads"
TEMP_IMAGE      = "/tmp/temp_image.jpg"
DB_PATH         = "/tmp/scan_history.db"
EXPORT_DIR      = "/tmp/exports"

# ─────────────────────────────────────────
# DATABASE  (v1.8 – Scan History)
# ─────────────────────────────────────────

def db_connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def db_init():
    with db_connect() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS scans (
                id          TEXT PRIMARY KEY,
                brand       TEXT,
                platform    TEXT,
                scanned_at  TEXT,
                total_posts INTEGER,
                high_risk   INTEGER,
                medium_risk INTEGER,
                low_risk    INTEGER,
                avg_score   REAL,
                logo_file   TEXT
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS detections (
                id          TEXT PRIMARY KEY,
                scan_id     TEXT,
                brand       TEXT,
                platform    TEXT,
                username    TEXT,
                post_url    TEXT,
                published   TEXT,
                ssim_score  REAL,
                hash_score  REAL,
                match_score REAL,
                risk        TEXT,
                description TEXT,
                FOREIGN KEY(scan_id) REFERENCES scans(id)
            )
        """)

db_init()

def db_save_scan(brand, platform, detections, logo_path):
    scan_id    = str(uuid.uuid4())
    total      = len(detections)
    high       = sum(1 for d in detections if d["risk"] == "High")
    medium     = sum(1 for d in detections if d["risk"] == "Medium")
    low        = sum(1 for d in detections if d["risk"] == "Low")
    avg        = round(sum(d["matchScore"] for d in detections) / total, 2) if total else 0
    logo_file  = os.path.basename(logo_path) if logo_path else "None"
    scanned_at = now_ist()

    with db_connect() as conn:
        conn.execute(
            "INSERT INTO scans VALUES (?,?,?,?,?,?,?,?,?,?)",
            (scan_id, brand, platform, scanned_at,
             total, high, medium, low, avg, logo_file)
        )
        for d in detections:
            conn.execute(
                "INSERT INTO detections VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                (str(uuid.uuid4()), scan_id, brand, platform,
                 d["username"], d["postUrl"], d["publishedDate"],
                 d["ssimScore"], d["hashScore"], d["matchScore"],
                 d["risk"], d["description"])
            )
    return scan_id

def db_get_scans(limit=50):
    with db_connect() as conn:
        return [dict(r) for r in conn.execute(
            "SELECT * FROM scans ORDER BY scanned_at DESC LIMIT ?", (limit,)
        )]

def db_get_scan(scan_id):
    with db_connect() as conn:
        scan = conn.execute(
            "SELECT * FROM scans WHERE id=?", (scan_id,)
        ).fetchone()
        dets = conn.execute(
            "SELECT * FROM detections WHERE scan_id=? ORDER BY match_score DESC",
            (scan_id,)
        ).fetchall()
    return dict(scan) if scan else None, [dict(d) for d in dets]

def db_repeat_offenders(brand=None, limit=10):
    """Accounts that appear in multiple scans — for v2.1.5 groundwork."""
    with db_connect() as conn:
        q = """
            SELECT username, platform, brand,
                   COUNT(*) as scan_count,
                   MAX(match_score) as max_score,
                   SUM(CASE WHEN risk='High' THEN 1 ELSE 0 END) as high_count
            FROM detections
            {where}
            GROUP BY username, platform, brand
            HAVING scan_count > 1
            ORDER BY high_count DESC, scan_count DESC
            LIMIT ?
        """
        if brand:
            rows = conn.execute(
                q.format(where="WHERE brand=?"), (brand, limit)
            ).fetchall()
        else:
            rows = conn.execute(
                q.format(where=""), (limit,)
            ).fetchall()
    return [dict(r) for r in rows]

# ─────────────────────────────────────────
# LOGO STATE
# ─────────────────────────────────────────

def get_logo_path():
    try:
        with open(LOGO_STATE_FILE) as f:
            state = json.load(f)
            p = state.get("logo_path")
            if p and os.path.exists(p):
                return p
    except Exception:
        pass
    return None

def set_logo_path(path):
    with open(LOGO_STATE_FILE, "w") as f:
        json.dump({"logo_path": path}, f)

# ─────────────────────────────────────────
# SCORING  (v1.6 — unchanged)
# ─────────────────────────────────────────

def compute_ssim_score(logo_gray, image_gray):
    try:
        resized = cv2.resize(image_gray, (logo_gray.shape[1], logo_gray.shape[0]))
        score, _ = ssim(logo_gray, resized, full=True)
        return round(max(score, 0) * 100, 2)
    except Exception:
        return 0.0

def compute_hash_score(logo_pil, image_pil):
    try:
        diff = imagehash.phash(logo_pil) - imagehash.phash(image_pil)
        return round(max(0, (64 - diff) / 64 * 100), 2)
    except Exception:
        return 0.0

def combined_score(s, h):
    return round(s * 0.35 + h * 0.65, 2)

def risk_level(score):
    if score >= 65:   return "High"
    elif score >= 35: return "Medium"
    else:             return "Low"

def compare_logo(logo_path, image_path):
    try:
        logo_cv  = cv2.imread(logo_path)
        image_cv = cv2.imread(image_path)
        if logo_cv is None or image_cv is None:
            return 0.0, 0.0, 0.0, "Low"
        logo_gray  = cv2.cvtColor(logo_cv,  cv2.COLOR_BGR2GRAY)
        image_gray = cv2.cvtColor(image_cv, cv2.COLOR_BGR2GRAY)
        logo_pil   = Image.open(logo_path).convert("RGB")
        image_pil  = Image.open(image_path).convert("RGB")
        s = compute_ssim_score(logo_gray, image_gray)
        h = compute_hash_score(logo_pil, image_pil)
        c = combined_score(s, h)
        return s, h, c, risk_level(c)
    except Exception as e:
        print("COMPARE ERROR:", e)
        return 0.0, 0.0, 0.0, "Low"

def fetch_og_image_via_apify(post_url, apify_token):
    """
    Use Apify's cheerio-scraper to fetch the Instagram post page
    and extract the og:image URL. Apify's servers have Instagram access.
    Returns an image URL string or None.
    """
    try:
        # Start a cheerio-scraper run to extract og:image from the post page
        run_resp = requests.post(
            "https://api.apify.com/v2/acts/apify~cheerio-scraper/runs",
            params={"token": apify_token},
            json={
                "startUrls": [{"url": post_url}],
                "pageFunction": """async function pageFunction(context) {
                    const { $ } = context;
                    const ogImage = $('meta[property="og:image"]').attr('content') || '';
                    return { ogImage };
                }""",
                "maxRequestsPerCrawl": 1,
                "maxConcurrency": 1,
            },
            timeout=15,
        )
        if run_resp.status_code not in (200, 201):
            print(f"OG-IMAGE RUN FAIL: {run_resp.status_code}")
            return None

        run_id = run_resp.json().get("data", {}).get("id")
        if not run_id:
            return None

        # Poll for completion (max 30s)
        import time
        for _ in range(15):
            time.sleep(2)
            status_resp = requests.get(
                f"https://api.apify.com/v2/acts/apify~cheerio-scraper/runs/{run_id}",
                params={"token": apify_token},
                timeout=10,
            )
            status = status_resp.json().get("data", {}).get("status", "")
            if status in ("SUCCEEDED", "FAILED", "ABORTED"):
                break

        if status != "SUCCEEDED":
            print(f"OG-IMAGE RUN STATUS: {status}")
            return None

        # Fetch dataset results
        dataset_id = status_resp.json()["data"]["defaultDatasetId"]
        items_resp = requests.get(
            f"https://api.apify.com/v2/datasets/{dataset_id}/items",
            params={"token": apify_token},
            timeout=10,
        )
        items = items_resp.json()
        if items and isinstance(items, list):
            og_image = items[0].get("ogImage", "")
            if og_image:
                print(f"OG-IMAGE FOUND: {og_image[:80]}")
                return og_image
    except Exception as e:
        print(f"OG-IMAGE ERROR: {e}")
    return None


def download_image(url, suffix="0", post_url="", is_instagram=False):
    """
    Download a post image with three fallback methods:
    1. Apify residential proxy  — needed for Instagram CDN (scontent-*.cdninstagram.com)
    2. Direct with browser UA   — works for Facebook / LinkedIn CDNs
    3. Skip and log
    Now that images[] is populated in Apify datasets, url is a real CDN image URL.
    """
    if not url:
        print(f"DOWNLOAD SKIP [{suffix}] — no URL")
        return None

    APIFY_TOKEN = os.getenv("APIFY_TOKEN", "")
    is_ig_cdn   = "cdninstagram.com" in url or "fbcdn.net" in url

    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        )
    }
    path = f"/tmp/post_img_{suffix}.jpg"

    # ── Method 1: Apify proxy (Instagram CDN requires residential IP) ──────
    if is_ig_cdn and APIFY_TOKEN:
        try:
            import urllib3; urllib3.disable_warnings()
            proxy_url = f"http://auto:{APIFY_TOKEN}@proxy.apify.com:8000"
            proxies   = {"http": proxy_url, "https": proxy_url}
            r = requests.get(url, timeout=30, headers=headers,
                             proxies=proxies, verify=False)
            if r.status_code == 200 and len(r.content) > 1000:
                with open(path, "wb") as f:
                    f.write(r.content)
                print(f"DOWNLOAD OK [proxy] [{suffix}] size={len(r.content)}")
                return path
            else:
                print(f"DOWNLOAD PROXY FAIL: HTTP {r.status_code} size={len(r.content)}")
        except Exception as e:
            print(f"DOWNLOAD PROXY ERROR: {e}")

    # ── Method 2: Direct (Facebook / LinkedIn CDNs) ───────────────────────
    try:
        r = requests.get(url, timeout=20, headers=headers)
        if r.status_code == 200 and len(r.content) > 1000:
            with open(path, "wb") as f:
                f.write(r.content)
            print(f"DOWNLOAD OK [direct] [{suffix}] size={len(r.content)}")
            return path
        else:
            print(f"DOWNLOAD DIRECT FAIL: HTTP {r.status_code} — {url[:80]}")
    except Exception as e:
        print(f"DOWNLOAD DIRECT ERROR: {e} — {url[:80]}")

    return None

def fmt_date(raw):
    try:
        dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        return dt.astimezone(ZoneInfo("Asia/Kolkata")).strftime("%d %b %Y, %I:%M %p IST")
    except Exception:
        return raw or "—"

def now_ist():
    return datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%d %b %Y, %I:%M %p IST")

def now_ts():
    return datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y%m%d_%H%M%S")

# ─────────────────────────────────────────
# BRAND CONFIG
# ─────────────────────────────────────────

BRANDS = {
    "ICICI":         {"instagram": "PATavqgLW7SsQhtH1", "facebook": "", "linkedin": ""},
    "Groww":         {"instagram": "E5X98iyKXmjN0WJBd", "facebook": "", "linkedin": ""},
    "Motilal Oswal": {"instagram": "DeK7uR4YafleAnnXG", "facebook": "", "linkedin": ""},
    "Tata Capital":  {"instagram": "985UpuFr2a5EnnSal", "facebook": "", "linkedin": ""},
    "Zerodha":       {"instagram": "aeZlx2bavuSezbBsd", "facebook": "", "linkedin": ""},
    "Upstox":        {"instagram": "JpjzemyvarleuBtRV", "facebook": "", "linkedin": ""},
    "SBI":           {"instagram": "sYb8wItw9rP48Hb88", "facebook": "", "linkedin": ""},
    "Anand Rathi":   {"instagram": "zzkpFEaglOX96YtKE", "facebook": "", "linkedin": ""},
}

PLATFORMS = ["instagram", "facebook", "linkedin"]
PLATFORM_DISPLAY = {"instagram": "Instagram", "facebook": "Facebook", "linkedin": "LinkedIn"}

RISK_COLOR = {"High": "#dc3545", "Medium": "#fd7e14", "Low": "#198754"}

def risk_badge(level):
    c = RISK_COLOR.get(level, "#333")
    return f'<span style="color:{c};font-weight:600">{level}</span>'

# ─────────────────────────────────────────
# SHARED FETCH + SCORE
# ─────────────────────────────────────────

def fetch_and_score(dataset_id, platform, brand, logo_path, apify_token):
    url  = f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={apify_token}"
    data = requests.get(url, timeout=30).json()
    if not isinstance(data, list):
        return [], 0

    print(f"SCAN START — brand={brand} platform={platform} posts={len(data)}")
    print(f"SCAN LOGO  — logo_path={logo_path!r} exists={logo_path and os.path.exists(logo_path)}")

    detections = []
    for idx, post in enumerate(data):
        try:
            caption = str(post.get("caption", "")).replace("\n", " ")
            raw_date = post.get("timestamp", "")
            published_date = fmt_date(raw_date)

            if platform == "instagram":
                # images[] is populated when extendOutputFunction is used in Apify
                # Falls back to displayUrl if images[] is empty
                apify_images = post.get("images") or []
                image_url = (
                    (apify_images[0] if apify_images and isinstance(apify_images[0], str) else "")
                    or (apify_images[0].get("url", "") if apify_images and isinstance(apify_images[0], dict) else "")
                    or post.get("displayUrl", "")
                    or post.get("imageUrl", "")
                )
                username = post.get("ownerUsername", "unknown")
                post_url = post.get("url", "#")
            elif platform == "facebook":
                image_url = (post.get("media") or [{}])[0].get("url", "") or post.get("imageUrl", "")
                username  = post.get("pageName", "") or post.get("authorName", "unknown")
                post_url  = post.get("url", "#") or post.get("postUrl", "#")
            else:
                image_url = post.get("image", "") or post.get("imageUrl", "")
                username  = post.get("authorName", "") or post.get("companyName", "unknown")
                post_url  = post.get("url", "#") or post.get("postUrl", "#")

            s = h = c = 0.0
            risk = "Low"
            if not logo_path:
                if idx == 0:
                    print("SCORE SKIP — no logo_path set")
            else:
                img_path = download_image(
                    url=image_url,
                    suffix=str(idx),
                    post_url=post_url,
                    is_instagram=(platform == "instagram"),
                )
                if img_path:
                    s, h, c, risk = compare_logo(logo_path, img_path)
                    print(f"SCORED [{idx}] {username} — ssim={s} hash={h} combined={c} risk={risk}")
                else:
                    print(f"SCORE SKIP [{idx}] {username} — all download methods failed")

            detections.append({
                "platform":      PLATFORM_DISPLAY[platform],
                "username":      username,
                "publishedDate": published_date,
                "postUrl":       post_url,
                "detectedBrand": brand,
                "ssimScore":     s,
                "hashScore":     h,
                "matchScore":    c,
                "risk":          risk,
                "description":   caption[:120] + "…" if len(caption) > 120 else caption,
            })
        except Exception as e:
            print(f"ITEM ERROR [{platform}] idx={idx}:", e)

    print(f"SCAN DONE  — {len(detections)} detections")
    return detections, len(data)

# ─────────────────────────────────────────
# SHARED PAGE CHROME
# ─────────────────────────────────────────

BASE_CSS = """
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:Arial,sans-serif;background:#f5f6fa;color:#222}
.page{max-width:1200px;margin:0 auto;padding:28px}
h1{font-size:20px;margin-bottom:4px}
.sub{font-size:13px;color:#666;margin-bottom:18px}
.banner{padding:10px 16px;border-radius:6px;font-size:13px;margin-bottom:14px}
.banner-ok{background:#d1fae5;border:1px solid #6ee7b7;color:#065f46}
.banner-warn{background:#fef3c7;border:1px solid #fcd34d;color:#92400e}
.banner-info{background:#e0f2fe;border:1px solid #7dd3fc;color:#075985}
.card{background:#fff;border-radius:8px;padding:22px;margin-bottom:18px;box-shadow:0 1px 4px rgba(0,0,0,.07)}
.card h2{font-size:15px;margin-bottom:8px;color:#333}
.card p{font-size:13px;color:#666;margin-bottom:12px;line-height:1.6}
.row{display:flex;flex-wrap:wrap;gap:10px;align-items:center}
select,input[type=file]{padding:9px;font-size:13px;border:1px solid #ccc;border-radius:4px}
select{min-width:200px}
.btn{padding:9px 18px;border-radius:4px;border:none;cursor:pointer;font-size:13px;
     text-decoration:none;display:inline-block;color:#fff;line-height:1}
.btn-blue{background:#0d6efd} .btn-blue:hover{background:#0b5ed7}
.btn-green{background:#198754} .btn-green:hover{background:#157347}
.btn-gray{background:#6c757d}  .btn-gray:hover{background:#5a6268}
.btn-red{background:#dc3545}   .btn-red:hover{background:#bb2d3b}
.scards{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:18px}
.scard{background:#fff;border-radius:6px;padding:14px 22px;min-width:120px;
       box-shadow:0 1px 3px rgba(0,0,0,.08);text-align:center}
.scard .num{font-size:28px;font-weight:700}
.scard .lbl{font-size:11px;color:#666;margin-top:3px}
.actions{display:flex;gap:10px;margin-bottom:16px;flex-wrap:wrap}
table{width:100%;border-collapse:collapse;background:#fff;
      box-shadow:0 1px 4px rgba(0,0,0,.07);font-size:13px}
th{background:#0d6efd;color:#fff;padding:10px;text-align:left}
td{padding:9px 10px;border-bottom:1px solid #eee;vertical-align:top}
tr:hover td{background:#f5f8ff}
a{color:#0d6efd;text-decoration:none}
.nav{display:flex;gap:16px;margin-bottom:22px;border-bottom:1px solid #e0e0e0;padding-bottom:12px;flex-wrap:wrap}
.nav a{font-size:13px;color:#555;text-decoration:none;padding-bottom:2px}
.nav a.active{color:#0d6efd;border-bottom:2px solid #0d6efd;font-weight:600}
.step-num{display:inline-block;background:#0d6efd;color:#fff;border-radius:50%;
          width:20px;height:20px;text-align:center;line-height:20px;font-size:11px;margin-right:6px}
"""

def nav_html(active="dashboard"):
    links = [
        ("dashboard", "/dashboard",  "Dashboard"),
        ("history",   "/history",    "Scan History"),
        ("offenders", "/offenders",  "Repeat Offenders"),
    ]
    parts = []
    for key, href, label in links:
        cls = 'class="active"' if key == active else ""
        parts.append(f'<a href="{href}" {cls}>{label}</a>')
    return '<div class="nav">' + "".join(parts) + "</div>"

# ─────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────

@app.get("/")
def home():
    return {"status": "AI Visual Threat Monitoring v1.9"}

# ── Debug ────────────────────────────────

@app.get("/debug-post")
def debug_post(brand: str = "ICICI", platform: str = "instagram"):
    """Shows raw Apify fields for the first post — use this to find the image URL field."""
    APIFY_TOKEN = os.getenv("APIFY_TOKEN")
    if not APIFY_TOKEN:
        return {"error": "Missing APIFY_TOKEN"}
    dataset_id = BRANDS.get(brand, {}).get(platform, "")
    if not dataset_id:
        return {"error": "No dataset configured"}
    url  = f"https://api.apify.com/v2/datasets/{dataset_id}/items?token={APIFY_TOKEN}&limit=1"
    data = requests.get(url, timeout=20).json()
    if not data:
        return {"error": "Empty dataset"}
    post = data[0]
    # Show all fields and their types/preview values
    summary = {}
    for k, v in post.items():
        if isinstance(v, str):
            summary[k] = v[:120]
        elif isinstance(v, list):
            summary[k] = f"list[{len(v)}]: {str(v[:2])[:200]}"
        elif isinstance(v, dict):
            summary[k] = f"dict keys: {list(v.keys())[:10]}"
        else:
            summary[k] = repr(v)[:120]
    return {"brand": brand, "platform": platform, "first_post_fields": summary}

@app.get("/debug")
def debug():
    logo_path = get_logo_path()
    tmp_files = os.listdir("/tmp") if os.path.exists("/tmp") else []
    upload_files = os.listdir("/tmp/uploads") if os.path.exists("/tmp/uploads") else []
    logo_size = os.path.getsize(logo_path) if logo_path and os.path.exists(logo_path) else 0
    return {
        "logo_path":       logo_path,
        "logo_exists":     bool(logo_path and os.path.exists(logo_path)),
        "logo_size_bytes": logo_size,
        "logo_state_file_exists": os.path.exists(LOGO_STATE_FILE),
        "logo_state_contents": open(LOGO_STATE_FILE).read() if os.path.exists(LOGO_STATE_FILE) else None,
        "tmp_uploads":     upload_files,
        "tmp_files":       [f for f in tmp_files if not f.startswith(".")],
        "worker_pid":      os.getpid(),
    }

# ── Upload ──────────────────────────────

@app.post("/upload")
async def upload_logo(file: UploadFile = File(...)):
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    filepath = os.path.join(UPLOAD_DIR, file.filename)
    with open(filepath, "wb") as buf:
        buf.write(await file.read())
    set_logo_path(filepath)
    return {"message": "Logo uploaded", "file": filepath, "status": "ready_to_scan"}

@app.get("/logo-status")
def logo_status():
    p = get_logo_path()
    return {"uploaded": bool(p), "filename": os.path.basename(p) if p else None}

# ── Dashboard ───────────────────────────

@app.get("/dashboard", response_class=HTMLResponse)
def dashboard():
    logo_path = get_logo_path()
    logo_filename = os.path.basename(logo_path) if logo_path else None

    logo_banner = (
        f'<div class="banner banner-ok">✅ Logo ready: <strong>{logo_filename}</strong></div>'
        if logo_path else
        '<div class="banner banner-warn">⚠️ No logo uploaded yet — upload below before scanning.</div>'
    )

    brand_opts    = "".join(f'<option value="{b}">{b}</option>' for b in BRANDS)
    platform_opts = "".join(f'<option value="{p}">{PLATFORM_DISPLAY[p]}</option>' for p in PLATFORMS)

    recent = db_get_scans(limit=5)
    recent_rows = ""
    for s in recent:
        recent_rows += f"""
        <tr>
          <td>{s['scanned_at']}</td>
          <td><strong>{s['brand']}</strong></td>
          <td>{s['platform']}</td>
          <td>{s['total_posts']}</td>
          <td style="color:#dc3545;font-weight:600">{s['high_risk']}</td>
          <td style="color:#fd7e14;font-weight:600">{s['medium_risk']}</td>
          <td>{s['avg_score']}%</td>
          <td><a href="/history/{s['id']}">View</a></td>
        </tr>"""

    recent_table = f"""
    <div class="card">
      <h2>Recent Scans</h2>
      <table>
        <tr><th>Date</th><th>Brand</th><th>Platform</th><th>Posts</th>
            <th>High</th><th>Medium</th><th>Avg Score</th><th></th></tr>
        {recent_rows if recent_rows else '<tr><td colspan="8" style="color:#999;padding:16px">No scans yet</td></tr>'}
      </table>
    </div>""" if recent else ""

    return f"""<!DOCTYPE html><html><head><title>Brand Monitoring</title>
<style>{BASE_CSS}</style></head><body><div class="page">
<h1>🛡️ AI Visual Threat Monitoring</h1>
<p class="sub">Brand logo comparison across Instagram, Facebook and LinkedIn</p>
{nav_html("dashboard")}
{logo_banner}

<div class="card">
  <h2><span class="step-num">1</span> Upload Brand Logo</h2>
  <p>Upload the official logo you want to compare against social media posts. PNG with transparent background works best.</p>
  <div class="row">
    <input type="file" id="logoFile" accept="image/*">
    <button class="btn btn-green" onclick="uploadLogo()">Upload Logo</button>
  </div>
  <div id="upload-result" style="margin-top:10px;font-size:13px"></div>
</div>

<div class="card">
  <h2><span class="step-num">2</span> Scan a Brand</h2>
  <p>Pick a brand and platform. The system downloads every post image, compares it against your logo using SSIM + ImageHash, and shows the match score and risk level.</p>
  <form action="/scan" method="get">
    <div class="row">
      <select name="brand">{brand_opts}</select>
      <select name="platform">{platform_opts}</select>
      <button type="submit" class="btn btn-blue">Scan Brand</button>
    </div>
  </form>
</div>

{recent_table}

</div>
<script>
async function uploadLogo() {{
  const f = document.getElementById('logoFile');
  const r = document.getElementById('upload-result');
  if (!f.files.length) {{ r.innerHTML='⚠️ Select a file first.'; return; }}
  r.innerHTML = 'Uploading…';
  const fd = new FormData();
  fd.append('file', f.files[0]);
  try {{
    const res  = await fetch('/upload', {{method:'POST',body:fd}});
    const data = await res.json();
    if (data.status === 'ready_to_scan') {{
      r.innerHTML = '✅ Uploaded: <strong>' + f.files[0].name + '</strong>';
      setTimeout(() => location.reload(), 1000);
    }} else {{ r.innerHTML = '❌ ' + JSON.stringify(data); }}
  }} catch(e) {{ r.innerHTML = '❌ ' + e.message; }}
}}
</script>
</body></html>"""

# ── Scan ────────────────────────────────

@app.get("/scan", response_class=HTMLResponse)
def scan(brand: str, platform: str = "instagram"):
    APIFY_TOKEN = os.getenv("APIFY_TOKEN")
    if not APIFY_TOKEN:
        return "<h2>Missing APIFY_TOKEN</h2>"
    if brand not in BRANDS:
        return f"<h2>Unknown brand: {brand}</h2>"
    if platform not in PLATFORMS:
        return f"<h2>Unknown platform: {platform}</h2>"

    dataset_id = BRANDS[brand].get(platform, "")
    if not dataset_id:
        return f"""<!DOCTYPE html><html><head><style>{BASE_CSS}</style></head><body><div class="page">
        <div class="banner banner-warn">No {PLATFORM_DISPLAY[platform]} dataset configured for {brand} yet.
        Add the Apify dataset ID to the BRANDS config in main.py.</div>
        <a class="btn btn-gray" href="/dashboard">← Dashboard</a></div></body></html>"""

    logo_path = get_logo_path()

    try:
        detections, total_records = fetch_and_score(dataset_id, platform, brand, logo_path, APIFY_TOKEN)
        detections.sort(key=lambda x: x["matchScore"], reverse=True)

        # Save to history
        scan_id = db_save_scan(brand, platform, detections, logo_path)

        high   = sum(1 for d in detections if d["risk"] == "High")
        medium = sum(1 for d in detections if d["risk"] == "Medium")
        low    = sum(1 for d in detections if d["risk"] == "Low")
        avg    = round(sum(d["matchScore"] for d in detections) / len(detections), 2) if detections else 0

        logo_banner = (
            f'<div class="banner banner-ok">✅ Logo: <strong>{os.path.basename(logo_path)}</strong>'
            f' &nbsp;·&nbsp; SSIM × 35% + ImageHash × 65%</div>'
            if logo_path else
            '<div class="banner banner-warn">⚠️ No logo — all scores are 0%. '
            '<a href="/dashboard">Upload a logo</a> first.</div>'
        )

        rows = ""
        for d in detections:
            rows += f"""<tr>
              <td>{d['platform']}</td>
              <td><strong>{d['username']}</strong></td>
              <td style="white-space:nowrap">{d['publishedDate']}</td>
              <td><span style="color:#0d6efd;font-weight:600">{d['detectedBrand']}</span></td>
              <td style="font-size:12px">{d['description']}</td>
              <td style="font-size:12px;color:#666">SSIM: {d['ssimScore']}%<br>Hash: {d['hashScore']}%</td>
              <td style="font-weight:700">{d['matchScore']}%</td>
              <td>{risk_badge(d['risk'])}</td>
              <td><a href="{d['postUrl']}" target="_blank">View ↗</a></td>
            </tr>"""

        # v1.7 – risk distribution chart data for inline Chart.js
        chart_data = json.dumps([high, medium, low])

        return f"""<!DOCTYPE html><html><head>
<title>{brand} — {PLATFORM_DISPLAY[platform]}</title>
<style>{BASE_CSS}
.chart-wrap{{background:#fff;border-radius:8px;padding:20px;margin-bottom:18px;
             box-shadow:0 1px 4px rgba(0,0,0,.07);max-width:420px}}
</style>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
</head><body><div class="page">
<h1>🛡️ {brand} — {PLATFORM_DISPLAY[platform]}</h1>
<p class="sub">Scan ID: {scan_id} &nbsp;·&nbsp; {now_ist()}</p>
{nav_html()}
{logo_banner}

<div class="scards">
  <div class="scard"><div class="num">{len(detections)}</div><div class="lbl">Total Posts</div></div>
  <div class="scard"><div class="num" style="color:#dc3545">{high}</div><div class="lbl">High Risk</div></div>
  <div class="scard"><div class="num" style="color:#fd7e14">{medium}</div><div class="lbl">Medium Risk</div></div>
  <div class="scard"><div class="num" style="color:#198754">{low}</div><div class="lbl">Low Risk</div></div>
  <div class="scard"><div class="num">{avg}%</div><div class="lbl">Avg Score</div></div>
</div>

<div class="chart-wrap">
  <canvas id="riskChart" height="180"></canvas>
</div>

<div class="actions">
  <a class="btn btn-gray" href="/dashboard">← Dashboard</a>
  <a class="btn btn-gray" href="/history">Scan History</a>
  <a class="btn btn-green" href="/export?brand={brand}&platform={platform}&scan_id={scan_id}">⬇ Export Excel</a>
</div>

<table>
  <tr><th>Platform</th><th>Username</th><th>Published</th><th>Brand</th>
      <th>Description</th><th>Score Breakdown</th><th>Match Score</th><th>Risk</th><th>Post</th></tr>
  {rows}
</table>
</div>
<script>
new Chart(document.getElementById('riskChart'), {{
  type: 'bar',
  data: {{
    labels: ['High Risk', 'Medium Risk', 'Low Risk'],
    datasets: [{{
      label: 'Posts',
      data: {chart_data},
      backgroundColor: ['#dc3545','#fd7e14','#198754'],
      borderRadius: 4,
      borderSkipped: false,
    }}]
  }},
  options: {{
    responsive: true,
    plugins: {{
      legend: {{ display: false }},
      title: {{ display: true, text: 'Risk Distribution', font: {{ size: 13 }} }}
    }},
    scales: {{ y: {{ beginAtZero: true, ticks: {{ stepSize: 1 }} }} }}
  }}
}});
</script>
</body></html>"""

    except Exception as e:
        print("SCAN ERROR:", e)
        return f"<h2>Error</h2><p>{e}</p>"

# ── Scan History List ────────────────────

@app.get("/history", response_class=HTMLResponse)
def history_list():
    scans = db_get_scans(limit=100)

    rows = ""
    for s in scans:
        high_style = 'style="color:#dc3545;font-weight:600"' if s["high_risk"] > 0 else ""
        rows += f"""<tr>
          <td style="white-space:nowrap">{s['scanned_at']}</td>
          <td><strong>{s['brand']}</strong></td>
          <td>{s['platform']}</td>
          <td>{s['total_posts']}</td>
          <td {high_style}>{s['high_risk']}</td>
          <td style="color:#fd7e14;font-weight:600">{s['medium_risk']}</td>
          <td style="color:#198754">{s['low_risk']}</td>
          <td>{s['avg_score']}%</td>
          <td>{s['logo_file']}</td>
          <td><a href="/history/{s['id']}">View</a></td>
        </tr>"""

    empty = '<tr><td colspan="10" style="color:#999;padding:20px;text-align:center">No scans yet. Run your first scan from the Dashboard.</td></tr>'

    return f"""<!DOCTYPE html><html><head><title>Scan History</title>
<style>{BASE_CSS}</style></head><body><div class="page">
<h1>🛡️ Scan History</h1>
<p class="sub">{len(scans)} scans recorded</p>
{nav_html("history")}
<div class="actions"><a class="btn btn-gray" href="/dashboard">← Dashboard</a></div>
<table>
  <tr><th>Date (IST)</th><th>Brand</th><th>Platform</th><th>Posts</th>
      <th>High</th><th>Medium</th><th>Low</th><th>Avg Score</th><th>Logo</th><th></th></tr>
  {rows if rows else empty}
</table>
</div></body></html>"""

# ── Scan History Detail ──────────────────

@app.get("/history/{scan_id}", response_class=HTMLResponse)
def history_detail(scan_id: str):
    scan, detections = db_get_scan(scan_id)
    if not scan:
        return "<h2>Scan not found</h2>"

    high   = scan["high_risk"]
    medium = scan["medium_risk"]
    low    = scan["low_risk"]
    chart_data = json.dumps([high, medium, low])

    rows = ""
    for d in detections:
        rows += f"""<tr>
          <td><strong>{d['username']}</strong></td>
          <td style="white-space:nowrap">{d['published']}</td>
          <td style="font-size:12px;color:#666">SSIM: {d['ssim_score']}%<br>Hash: {d['hash_score']}%</td>
          <td style="font-weight:700">{d['match_score']}%</td>
          <td>{risk_badge(d['risk'])}</td>
          <td style="font-size:12px">{d['description']}</td>
          <td><a href="{d['post_url']}" target="_blank">View ↗</a></td>
        </tr>"""

    return f"""<!DOCTYPE html><html><head>
<title>Scan {scan_id[:8]}</title>
<style>{BASE_CSS}
.chart-wrap{{background:#fff;border-radius:8px;padding:20px;margin-bottom:18px;
             box-shadow:0 1px 4px rgba(0,0,0,.07);max-width:400px}}
</style>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
</head><body><div class="page">
<h1>🛡️ {scan['brand']} — {scan['platform']}</h1>
<p class="sub">Scan on {scan['scanned_at']} &nbsp;·&nbsp; Logo: {scan['logo_file']}</p>
{nav_html("history")}

<div class="scards">
  <div class="scard"><div class="num">{scan['total_posts']}</div><div class="lbl">Total Posts</div></div>
  <div class="scard"><div class="num" style="color:#dc3545">{high}</div><div class="lbl">High Risk</div></div>
  <div class="scard"><div class="num" style="color:#fd7e14">{medium}</div><div class="lbl">Medium Risk</div></div>
  <div class="scard"><div class="num" style="color:#198754">{low}</div><div class="lbl">Low Risk</div></div>
  <div class="scard"><div class="num">{scan['avg_score']}%</div><div class="lbl">Avg Score</div></div>
</div>

<div class="chart-wrap">
  <canvas id="riskChart" height="180"></canvas>
</div>

<div class="actions">
  <a class="btn btn-gray" href="/history">← History</a>
  <a class="btn btn-green" href="/export?brand={scan['brand']}&platform={scan['platform'].lower()}&scan_id={scan_id}">⬇ Export Excel</a>
</div>

<table>
  <tr><th>Username</th><th>Published</th><th>Score Breakdown</th>
      <th>Match Score</th><th>Risk</th><th>Description</th><th>Post</th></tr>
  {rows}
</table>
</div>
<script>
new Chart(document.getElementById('riskChart'), {{
  type: 'doughnut',
  data: {{
    labels: ['High','Medium','Low'],
    datasets:[{{data:{chart_data},backgroundColor:['#dc3545','#fd7e14','#198754'],borderWidth:2}}]
  }},
  options:{{responsive:true,plugins:{{legend:{{position:'right'}},
    title:{{display:true,text:'Risk Distribution',font:{{size:13}}}}}}}}
}});
</script>
</body></html>"""

# ── Repeat Offenders (v2.1.5 groundwork) ─

@app.get("/offenders", response_class=HTMLResponse)
def repeat_offenders(brand: str = ""):
    offenders = db_repeat_offenders(brand=brand or None, limit=50)

    brand_opts = '<option value="">All Brands</option>' + "".join(
        f'<option value="{b}" {"selected" if b==brand else ""}>{b}</option>'
        for b in BRANDS
    )

    rows = ""
    for o in offenders:
        rows += f"""<tr>
          <td><strong>{o['username']}</strong></td>
          <td>{o['platform']}</td>
          <td>{o['brand']}</td>
          <td style="font-weight:700">{o['scan_count']}</td>
          <td style="font-weight:700;color:#dc3545">{o['high_count']}</td>
          <td>{o['max_score']}%</td>
        </tr>"""

    empty = '<tr><td colspan="6" style="color:#999;padding:20px;text-align:center">No repeat offenders found yet. Accounts appear here after they show up in 2+ scans.</td></tr>'

    return f"""<!DOCTYPE html><html><head><title>Repeat Offenders</title>
<style>{BASE_CSS}</style></head><body><div class="page">
<h1>🛡️ Repeat Offenders</h1>
<p class="sub">Accounts detected across multiple scans — highest risk first</p>
{nav_html("offenders")}

<div class="card">
  <form method="get" style="display:flex;gap:10px;align-items:center">
    <select name="brand">{brand_opts}</select>
    <button type="submit" class="btn btn-blue">Filter</button>
  </form>
</div>

<div class="banner banner-info">
  Accounts listed here have appeared in <strong>2 or more scans</strong>.
  High-risk repeat accounts are the strongest signal of brand abuse or impersonation.
</div>

<table>
  <tr><th>Username</th><th>Platform</th><th>Brand</th>
      <th>Times Seen</th><th>High Risk Hits</th><th>Highest Score</th></tr>
  {rows if rows else empty}
</table>
</div></body></html>"""

# ── Export Excel  (v1.9 – Advanced) ─────

@app.get("/export")
def export_excel(brand: str, platform: str = "instagram", scan_id: str = ""):

    APIFY_TOKEN = os.getenv("APIFY_TOKEN")
    if not APIFY_TOKEN:
        return {"error": "Missing APIFY_TOKEN"}
    if brand not in BRANDS:
        return {"error": f"Unknown brand: {brand}"}

    logo_path = get_logo_path()

    # Use saved detections if scan_id provided, else re-fetch
    if scan_id:
        scan, saved_dets = db_get_scan(scan_id)
        if scan and saved_dets:
            detections = [{
                "platform":      scan["platform"],
                "username":      d["username"],
                "publishedDate": d["published"],
                "postUrl":       d["post_url"],
                "detectedBrand": brand,
                "ssimScore":     d["ssim_score"],
                "hashScore":     d["hash_score"],
                "matchScore":    d["match_score"],
                "risk":          d["risk"],
                "description":   d["description"],
            } for d in saved_dets]
        else:
            return {"error": "Scan ID not found"}
    else:
        dataset_id = BRANDS[brand].get(platform, "")
        if not dataset_id:
            return {"error": f"No {platform} dataset for {brand}"}
        detections, _ = fetch_and_score(dataset_id, platform, brand, logo_path, APIFY_TOKEN)
        detections.sort(key=lambda x: x["matchScore"], reverse=True)

    high   = sum(1 for d in detections if d["risk"] == "High")
    medium = sum(1 for d in detections if d["risk"] == "Medium")
    low    = sum(1 for d in detections if d["risk"] == "Low")
    avg    = round(sum(d["matchScore"] for d in detections) / len(detections), 2) if detections else 0

    os.makedirs(EXPORT_DIR, exist_ok=True)
    ts       = now_ts()
    filename = f"{EXPORT_DIR}/{brand}_{platform}_{ts}.xlsx"

    # ── Build workbook with openpyxl directly for full styling ──
    from openpyxl import Workbook
    wb = Workbook()

    # Colour constants
    BLUE_FILL   = PatternFill("solid", fgColor="0D6EFD")
    RED_FILL    = PatternFill("solid", fgColor="DC3545")
    ORANGE_FILL = PatternFill("solid", fgColor="FD7E14")
    GREEN_FILL  = PatternFill("solid", fgColor="198754")
    GRAY_FILL   = PatternFill("solid", fgColor="F8F9FA")
    WHITE_FONT  = Font(color="FFFFFF", bold=True, size=11)
    BOLD        = Font(bold=True)
    THIN_BORDER = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    def style_header_row(ws, row_num, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill   = BLUE_FILL
            cell.font   = WHITE_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_data_cell(cell, center=False):
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=True
        )

    # ── Sheet 1: Executive Summary ─────────────────────────────
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells("A1:F1")
    title_cell = ws1["A1"]
    title_cell.value     = f"Brand Monitoring Report — {brand} ({PLATFORM_DISPLAY[platform]})"
    title_cell.font      = Font(bold=True, size=14, color="0D6EFD")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    summary_data = [
        ("Brand",              brand),
        ("Platform",           PLATFORM_DISPLAY[platform]),
        ("Report Date (IST)",  now_ist()),
        ("Logo Used",          os.path.basename(logo_path) if logo_path else "None"),
        ("Scoring Method",     "SSIM 35% + ImageHash 65%"),
        ("",                   ""),
        ("Total Posts Scanned",  len(detections)),
        ("High Risk Posts",      high),
        ("Medium Risk Posts",    medium),
        ("Low Risk Posts",       low),
        ("Average Match Score",  f"{avg}%"),
    ]

    for i, (label, value) in enumerate(summary_data, start=3):
        lc = ws1.cell(row=i, column=1, value=label)
        vc = ws1.cell(row=i, column=2, value=value)
        if label:
            lc.font   = BOLD
            lc.fill   = GRAY_FILL
            lc.border = THIN_BORDER
            vc.border = THIN_BORDER
            lc.alignment = Alignment(vertical="center")
            vc.alignment = Alignment(vertical="center")
            if label == "High Risk Posts":
                vc.fill = PatternFill("solid", fgColor="FFE0E3")
                vc.font = Font(color="DC3545", bold=True)
            elif label == "Medium Risk Posts":
                vc.fill = PatternFill("solid", fgColor="FFF0DB")
                vc.font = Font(color="FD7E14", bold=True)

    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 42

    # ── Sheet 2: All Detections ────────────────────────────────
    ws2 = wb.create_sheet("All Detections")
    ws2.sheet_view.showGridLines = False

    headers2 = ["Platform","Username","Published Date","Post URL",
                 "Detected Brand","SSIM Score (%)","Hash Score (%)","Match Score (%)","Risk","Description"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header_row(ws2, 1, len(headers2))
    ws2.row_dimensions[1].height = 28

    for row_i, d in enumerate(detections, start=2):
        vals = [d["platform"], d["username"], d["publishedDate"], d["postUrl"],
                d["detectedBrand"], d["ssimScore"], d["hashScore"], d["matchScore"],
                d["risk"], d["description"]]
        for col_i, val in enumerate(vals, 1):
            cell = ws2.cell(row=row_i, column=col_i, value=val)
            style_data_cell(cell, center=(col_i in [1,5,6,7,8,9]))
            ws2.row_dimensions[row_i].height = 20
            if col_i == 9:
                if val == "High":
                    cell.font = Font(color="DC3545", bold=True)
                    cell.fill = PatternFill("solid", fgColor="FFE0E3")
                elif val == "Medium":
                    cell.font = Font(color="C05C00", bold=True)
                    cell.fill = PatternFill("solid", fgColor="FFF0DB")
                else:
                    cell.font = Font(color="155724", bold=True)
                    cell.fill = PatternFill("solid", fgColor="D1FAE5")

    col_widths2 = [12,22,24,50,16,14,14,16,10,60]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[ws2.cell(1,i).column_letter].width = w

    # ── Sheet 3: High Risk Only ────────────────────────────────
    ws3 = wb.create_sheet("High Risk Posts")
    ws3.sheet_view.showGridLines = False

    headers3 = ["Username","Published Date","Post URL","Match Score (%)","SSIM (%)","Hash (%)","Description"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
        ws3.cell(row=1, column=col).fill   = RED_FILL
        ws3.cell(row=1, column=col).font   = WHITE_FONT
        ws3.cell(row=1, column=col).border = THIN_BORDER
        ws3.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    high_dets = [d for d in detections if d["risk"] == "High"]
    for row_i, d in enumerate(high_dets, start=2):
        vals = [d["username"], d["publishedDate"], d["postUrl"],
                d["matchScore"], d["ssimScore"], d["hashScore"], d["description"]]
        for col_i, val in enumerate(vals, 1):
            cell = ws3.cell(row=row_i, column=col_i, value=val)
            style_data_cell(cell, center=(col_i in [4,5,6]))
            ws3.row_dimensions[row_i].height = 20

    for i, w in enumerate([22,24,50,16,12,12,60], 1):
        ws3.column_dimensions[ws3.cell(1,i).column_letter].width = w

    # ── Sheet 4: Risk Chart (openpyxl bar chart) ───────────────
    ws4 = wb.create_sheet("Risk Chart")
    ws4.sheet_view.showGridLines = False
    ws4["A1"] = "Risk Level"; ws4["B1"] = "Count"
    ws4["A2"] = "High";   ws4["B2"] = high
    ws4["A3"] = "Medium"; ws4["B3"] = medium
    ws4["A4"] = "Low";    ws4["B4"] = low

    chart = BarChart()
    chart.type    = "col"
    chart.title   = f"Risk Distribution — {brand}"
    chart.y_axis.title = "Posts"
    chart.x_axis.title = "Risk Level"
    chart.width   = 20
    chart.height  = 14
    chart.shape   = 4

    data_ref  = Reference(ws4, min_col=2, min_row=1, max_row=4)
    cats_ref  = Reference(ws4, min_col=1, min_row=2, max_row=4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "0D6EFD"

    # Colour each bar individually
    for idx, color in enumerate(["DC3545","FD7E14","198754"]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        chart.series[0].dPt.append(pt)

    ws4.add_chart(chart, "D2")

    wb.save(filename)

    return FileResponse(
        path=filename,
        filename=f"{brand}_{platform}_{ts}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
